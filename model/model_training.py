import numpy as np
import logging
import matplotlib.pyplot as plt
from keras.api import Sequential, layers
from sklearn.preprocessing import MinMaxScaler
from sklearn.model_selection import KFold
from indicators.technical_indicators import calculate_rsi, calculate_macd, calculate_bollinger_bands, calculate_stochastic, calculate_ema, calculate_atr, calculate_cci, calculate_obv
from bi_api.data_extraction import get_historical_data
import tensorflow as tf
from tensorflow import keras
import keras_tuner as kt
from keras.regularizers import l1_l2
import pandas as pd
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

def train_lstm_model(symbol='BTCUSDT', look_back=60, epochs=50, batch_size=32, n_splits=2):
    logger.info(f"Завантаження даних для {symbol}.")
    data = get_historical_data(symbol)

    if data is None or data.empty:
        logger.error(f"Дані для {symbol} не були завантажені.")
        return None, None  # Повертаємо None для обох очікуваних значень

    # Розрахунок технічних індикаторів
    logger.info(f"Розрахунок технічних індикаторів для {symbol}.")
    data['RSI'] = calculate_rsi(data)
    data['MACD'], data['MACD_Signal'] = calculate_macd(data['close'])
    data['Upper_Band'], data['Lower_Band'] = calculate_bollinger_bands(data)
    data['Stoch'], data['Stoch_Signal'] = calculate_stochastic(data)
    data['EMA'] = calculate_ema(data)
    data['ATR'] = calculate_atr(data)
    data['CCI'] = calculate_cci(data)
    data['OBV'] = calculate_obv(data)

    data = data.dropna()

    if data.empty:
        logger.error(f"Після розрахунку індикаторів немає достатньо даних для {symbol}.")
        return None, None  # Повертаємо None для обох очікуваних значень

    logger.info("Масштабування даних.")
    scaler = MinMaxScaler(feature_range=(0, 1))
    scaled_data = scaler.fit_transform(data[['close', 'RSI', 'MACD', 'MACD_Signal', 'EMA', 'ATR', 'CCI', 'OBV']].values)

    logger.info(f"Форма масштабованих даних: {scaled_data.shape}")

    logger.info("Формування навчальних даних для моделі LSTM.")
    X, y = [], []
    for i in range(look_back, len(scaled_data)):
        X.append(scaled_data[i - look_back:i])
        y.append(scaled_data[i, 0])

    X, y = np.array(X), np.array(y)

    if X.shape[0] == 0:
        logger.error("Недостатньо даних для тренування.")
        return None, None  # Повертаємо None для обох очікуваних значень

    logger.info(f"Форма X: {X.shape}, форма y: {y.shape}")

    # Тренування моделі через крос-валідацію
    logger.info("Початок крос-валідації та тренування.")
    histories, filename = cross_validate_lstm(X, y, epochs, batch_size, n_splits)

    # Якщо немає історії, повертаємо None
    if not histories:
        logger.error("Крос-валідація не повернула жодної історії тренування.")
        return None, None

    return histories, filename  # Повертаємо історію тренувань та назву файлу Excel


def cross_validate_lstm(X, y, epochs, batch_size, n_splits):
    kfold = KFold(n_splits=n_splits, shuffle=True)
    fold = 1
    histories = []
    filename = 'training_history.xlsx'  # Назва файлу для збереження історії тренування

    for train_index, val_index in kfold.split(X):
        X_train, X_val = X[train_index], X[val_index]
        y_train, y_val = y[train_index], y[val_index]

        # Побудова моделі
        model = build_model_with_attention(X_train.shape[1:])

        # Тренування моделі
        history = model.fit(X_train, y_train, epochs=epochs, batch_size=batch_size, validation_data=(X_val, y_val))

        if history is not None:
            histories.append(history)  # Збереження історії тренування для кожного фолду
            export_training_history(history, fold, filename)  # Експорт результатів тренування у Excel

            logger.info(f"Модель для Fold {fold} завершена.")
        else:
            logger.error(f"Помилка тренування для Fold {fold}. Історія не збережена.")

        fold += 1

    return histories, filename  # Повертаємо всі історії тренувань та файл Excel

def build_model_with_attention(input_shape):
    inputs = layers.Input(shape=input_shape)

    # Bidirectional LSTM
    lstm_out = layers.Bidirectional(layers.LSTM(128, return_sequences=True))(inputs)

    # Застосування Attention
    attention_out = layers.Attention()([lstm_out, lstm_out])

    # Flatten після застосування Attention, щоб узгодити форму для Dense шару
    flatten = layers.Flatten()(attention_out)

    # Продовження побудови моделі
    dense = layers.Dense(64, activation='relu')(flatten)
    dropout = layers.Dropout(0.2)(dense)
    outputs = layers.Dense(1, activation='sigmoid')(dropout)

    # Вихідний шар з L1 та L2 регуляризацією
    outputs = layers.Dense(1, activation='sigmoid',
                           kernel_regularizer=l1_l2(l1=1e-5, l2=1e-4))(dropout)

    # Створення моделі
    model = keras.Model(inputs=inputs, outputs=outputs)

    # Компіляція моделі
    model.compile(optimizer='adam', loss='binary_crossentropy', metrics=['accuracy'])

    return model

# Гіперпараметричний пошук з Keras Tuner
def hyperparameter_search(X_train, y_train, epochs):
    def model_builder(hp):
        model = keras.Sequential()
        hp_units = hp.Int('units', min_value=32, max_value=128, step=16)
        model.add(keras.layers.LSTM(units=hp_units, input_shape=(X_train.shape[1], X_train.shape[2]), return_sequences=True))

        hp_dropout = hp.Float('dropout', min_value=0.0, max_value=0.5, step=0.1)
        model.add(keras.layers.Dropout(hp_dropout))

        model.add(keras.layers.Dense(units=1))
        model.compile(optimizer='adam', loss='mean_squared_error')
        return model

    tuner = kt.RandomSearch(
        model_builder,
        objective='val_loss',
        max_trials=10,
        executions_per_trial=3
    )

    tuner.search(X_train, y_train, epochs=epochs, validation_split=0.2)
    best_model = tuner.get_best_models(num_models=1)[0]
    return best_model

def export_training_history(history, fold_index, filename):
    history_df = pd.DataFrame(history.history)
    sheet_name = f'Fold_{fold_index + 1}'

    try:
        # Відкриваємо існуючий файл
        book = load_workbook(filename)

        # Якщо аркуш вже існує, видаляємо його
        if sheet_name in book.sheetnames:
            std = book[sheet_name]
            book.remove(std)

        # Використовуємо book для створення нових аркушів без необхідності встановлення 'book'
        with pd.ExcelWriter(filename, engine='openpyxl', mode='a') as writer:
            writer._book = book  # Використовуємо прихований атрибут _book
            history_df.to_excel(writer, sheet_name=sheet_name, index=False)
    except FileNotFoundError:
        # Якщо файл не існує, створюємо новий
        with pd.ExcelWriter(filename, engine='openpyxl', mode='w') as writer:
            history_df.to_excel(writer, sheet_name=sheet_name, index=False)

    print(f"Історію тренування для Fold {fold_index + 1} збережено до {filename}")

